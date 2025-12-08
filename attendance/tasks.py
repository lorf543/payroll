# attendance/tasks.py
import os
from datetime import datetime
from urllib.parse import urljoin
from tempfile import NamedTemporaryFile
from django.core.mail import EmailMessage

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.mail import send_mail
from django.conf import settings
from django.utils import timezone
from core.models import Employee
from .models import WorkDay


def generate_and_email_team_report(supervisor_id, date_from, date_to):
    """
    Genera un reporte de equipo en Excel y envía un correo con el archivo adjunto.
    """

    # Obtener supervisor
    supervisor = Employee.objects.get(id=supervisor_id)

    date_from_dt = datetime.strptime(date_from, "%Y-%m-%d").date()
    date_to_dt = datetime.strptime(date_to, "%Y-%m-%d").date()

    # Obtener miembros del equipo
    team_members = Employee.objects.filter(supervisor=supervisor)

    work_days = (
        WorkDay.objects.filter(employee__in=team_members, date__range=(date_from_dt, date_to_dt))
        .select_related("employee", "employee__user")
        .prefetch_related("sessions")
        .order_by("date", "employee__user__last_name")
    )

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Team Report"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    # Títulos
    ws.merge_cells("A1:Q1")
    ws["A1"] = f"Team Report - Supervisor: {supervisor.user.get_full_name()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:Q2")
    ws["A2"] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].alignment = center

    ws.append([])

    headers = [
        "Agent Name", "Date", "Attendance Status", "Week",
        "Time In",
        "Break 1-start", "Break 1-end", "Break1 Duration",
        "Break 2-start", "Break 2-end", "Break2 Duration",
        "Lunch start", "Lunch end", "Lunch Duration",
        "Time Out",
        "Total Hours",
        "Notes",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=4, column=col).font = bold
        ws.cell(row=4, column=col).alignment = center

    # Helpers
    def fmt_time(t):
        return t.strftime("%H:%M") if t else "-"

    def fmt_duration(s):
        if not s or not s.start_time or not s.end_time:
            return "-"
        delta = s.end_time - s.start_time
        minutes = delta.total_seconds() // 60
        seconds = int(delta.total_seconds() % 60)
        return f"{int(minutes):02d}:{seconds:02d}"

    def fmt_total(td):
        if td is None:
            return "0.00"
        return f"{td.total_seconds() / 3600:.2f}"

    # Filas
    for wd in work_days:
        sessions = wd.sessions.all()
        work_sessions = sessions.filter(session_type="work").order_by("start_time")
        break_sessions = sessions.filter(session_type="break").order_by("start_time")
        lunch_sessions = sessions.filter(session_type="lunch").order_by("start_time")

        time_in = work_sessions.first().start_time if work_sessions else None
        time_out = work_sessions.last().end_time if work_sessions else None
        break1 = break_sessions[0] if len(break_sessions) >= 1 else None
        break2 = break_sessions[1] if len(break_sessions) >= 2 else None
        lunch = lunch_sessions[0] if lunch_sessions else None

        ws.append([
            wd.employee.user.get_full_name(),
            wd.date.strftime("%m/%d/%Y"),
            wd.get_day_status(),
            wd.date.isocalendar()[1],
            fmt_time(time_in),
            fmt_time(break1.start_time if break1 else None),
            fmt_time(break1.end_time if break1 else None),
            fmt_duration(break1) if break1 else "-",
            fmt_time(break2.start_time if break2 else None),
            fmt_time(break2.end_time if break2 else None),
            fmt_duration(break2) if break2 else "-",
            fmt_time(lunch.start_time if lunch else None),
            fmt_time(lunch.end_time if lunch else None),
            fmt_duration(lunch) if lunch else "-",
            fmt_time(time_out),
            fmt_total(wd.total_work_time),
            wd.notes or "",
        ])

    # Ajustar ancho columnas
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2

    # Crear nombre del archivo
    filename = f"team_report_{supervisor.user.username}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Guardar en archivo temporal
    with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        wb.save(tmp_file.name)
        temp_path = tmp_file.name

    try:
        # Crear y enviar email con adjunto
        email = EmailMessage(
            subject="Team Report Ready",
            body=f"""Hi {supervisor.user.get_full_name()},

            Your team report for the period {date_from} to {date_to} is attached to this email.

            Report Details:
            - Team Members: {team_members.count()}
            - Total Records: {work_days.count()}
            - Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}

            Regards,
            Your System
            """,
            
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[supervisor.user.email],
        )

        # Adjuntar archivo
        with open(temp_path, 'rb') as f:
            email.attach(
                filename, 
                f.read(), 
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )

        email.send(fail_silently=False)

    finally:
        # Limpiar archivo temporal
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return filename