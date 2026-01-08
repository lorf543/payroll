from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseForbidden,HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import user_passes_test
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.contrib import messages
from django.contrib.sessions.models import Session
from django.db import transaction
from datetime import datetime, time, timedelta,date
from django.core.paginator import Paginator
from django.db.models import Q, Value
from django.db.models.functions import Concat
from django.contrib.auth.models import User
from django.contrib.auth import logout
from django.db import transaction
from django_q.tasks import async_task
from django.core.files.storage import default_storage
from django.http import FileResponse
from django.core.files.base import ContentFile


import os
import decimal
import openpyxl
import json

from django.views.generic import DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin



from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter



from core.models import Employee, Campaign
from .forms import EmployeeProfileForm,ActivitySessionForm, OccurrenceForm
from .models import WorkDay,ActivitySession, Occurrence
from .status_helpers import close_active_status
from .utility import *
from core.utils.payroll import get_effective_pay_rate
from .tasks import generate_and_email_team_report
# Create your views here.




def format_duration_simple(duration):
    """
    Acepta timedelta, segundos, minutos o strings numéricos y devuelve 'Xh YYm'
    """
    if not duration:
        return "0h 00m"

    try:
        # Si es timedelta → convertir a segundos
        if hasattr(duration, "total_seconds"):
            total_seconds = int(duration.total_seconds())

        # Si es string numérico → convertir
        elif isinstance(duration, str) and duration.isdigit():
            total_seconds = int(duration)

        # Si es número → asumir que viene en segundos grandes o minutos pequeños
        elif isinstance(duration, (int, float)):
            # Heurística: si es muy grande, probablemente ya está en segundos
            if duration > 3600:  # más de 1 hora en segundos
                total_seconds = int(duration)
            else:
                # probablemente son minutos → convertir a segundos
                total_seconds = int(duration) * 60
        else:
            return "0h 00m"

        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60

        if hours > 0:
            return f"{hours}h {minutes:02d}m"
        return f"{minutes}m"

    except Exception:
        return "0h 00m"


def force_logout_all_users(request):
    """
    Deletes all active sessions and properly closes all work sessions.
    """
    now = timezone.now()
    
    try:
        with transaction.atomic():
            # 1. ✅ Cerrar todas las ActivitySession activas
            active_sessions = ActivitySession.objects.filter(end_time__isnull=True)
            session_count = active_sessions.count()
            
            for session in active_sessions:
                session.end_time = now
                session.notes = f"{session.notes or ''}\nForcefully closed by system at {now}".strip()
                session.save(update_fields=['end_time', 'notes'])
            
            # 2. ✅ Actualizar WorkDays activos
            active_workdays = WorkDay.objects.filter(status='active')
            workday_count = active_workdays.count()
            
            for workday in active_workdays:
                workday.check_out = now
                workday.status = 'completed'
                workday.notes = f"{workday.notes or ''}\nForcefully closed by system at {now}".strip()
                workday.calculate_daily_totals()  # Recalcular métricas
                workday.save()
            
            # 3. ✅ Marcar todos los empleados como logout
            employee_count = Employee.objects.filter(is_logged_in=True).update(
                is_logged_in=False,
                last_logout=now
            )
            
            # 4. ✅ Eliminar sesiones de autenticación
            session_auth_count = Session.objects.all().count()
            Session.objects.all().delete()
            
            # Mensaje de confirmación
            message = (
                f"✅ Force logout completed at {now}:\n"
                f"• {session_count} active work sessions closed\n"
                f"• {workday_count} active workdays completed\n" 
                f"• {employee_count} employees logged out\n"
                f"• {session_auth_count} authentication sessions deleted"
            )
            
            print(f"[{datetime.now()}] {message}")
            
            return HttpResponse(message)
            
    except Exception as e:
        error_message = f"❌ Error during force logout: {str(e)}"
        print(f"[{datetime.now()}] {error_message}")
        return HttpResponse(error_message, status=500)


def is_supervisor(user):
    return user.is_staff or user.groups.filter(name='Supervisors').exists()

@login_required
def agent_dashboard(request):
    """
    Main dashboard view that populates all template information
    Uses WORK DAY logic instead of calendar day to handle midnight crossover
    """
    employee = get_object_or_404(Employee, user=request.user)
    
    # 🔥 CRITICAL: Get or create active work day with auto-close logic
    work_day = get_or_create_active_work_day(employee)
    
    # Get current active session
    current_session = work_day.get_active_session()
    
    # Get all sessions for this work day (history)
    history = work_day.sessions.all().order_by('start_time')
    
    # Force calculate daily totals
    calculate_daily_totals_manual(work_day)
    work_day.refresh_from_db()
    
    # Get campaign for break/lunch durations
    campaign = employee.current_campaign
    
    # Calculate all statistics
    daily_stats = calculate_daily_stats(work_day, employee)
    
    context = {
        'employee': employee,
        'work_day': work_day,
        'current_session': current_session,
        'history': history,
        'today': work_day.date,
        'campaign': campaign,
        'daily_stats': daily_stats,
    }
    
    return render(request, 'attendance/agent_dashboard.html', context)


def get_or_create_active_work_day(employee):
    """
    Get or create active work day for employee with smart auto-close logic.
    
    IMPORTANT: This allows work days to span past midnight BUT auto-closes
    old days that were forgotten.
    
    Logic:
    1. Find the most recent incomplete work day
    2. If it exists AND is "recent" (within 12 hours), use it
    3. If it's "old" (more than 12 hours), auto-close it and create new one
    4. If no incomplete day exists, create a new one for today
    
    This prevents the bug where a forgotten day accumulates hours from future days.
    """
    from django.db.models import Max
    
    now = timezone.now()
    today = now.date()
    
    # Find the most recent incomplete work day
    active_work_day = WorkDay.objects.filter(
        employee=employee,
        status__in=['active', 'regular_hours']  # Not completed
    ).order_by('-date').first()
    
    if active_work_day:
        # Check if this work day is "recent" or "old"
        # We consider it "recent" if:
        # 1. It's from today, OR
        # 2. It has an active session (employee is currently working), OR
        # 3. The last activity was within the last 12 hours
        
        is_today = active_work_day.date == today
        has_active_session = active_work_day.get_active_session() is not None
        
        # Get the last activity time (most recent session end or start)
        last_session = active_work_day.sessions.order_by('-start_time').first()
        
        if last_session:
            # Check the most recent timestamp from this session
            if last_session.end_time:
                last_activity = last_session.end_time
            else:
                # Active session - use start time
                last_activity = last_session.start_time
            
            # Calculate hours since last activity
            hours_since_activity = (now - last_activity).total_seconds() / 3600
            is_recent = hours_since_activity <= 12  # Within 12 hours
        else:
            # No sessions yet - check check_in time
            if active_work_day.check_in:
                hours_since_activity = (now - active_work_day.check_in).total_seconds() / 3600
                is_recent = hours_since_activity <= 12
            else:
                # No check_in either - it's old
                is_recent = False
        
        # Decision: Use this work day or auto-close it?
        if is_today or has_active_session or is_recent:
            # ✅ This is a valid active work day
            return active_work_day
        else:
            # ❌ This work day is old and forgotten - auto-close it
            auto_close_forgotten_workday(active_work_day)
            # Continue to create a new work day below
    
    # Create a new work day for today
    work_day, created = WorkDay.objects.get_or_create(
        employee=employee,
        date=today,
        defaults={'status': 'regular_hours'}
    )
    
    return work_day


def auto_close_forgotten_workday(work_day):
    """
    Auto-close a work day that was forgotten (employee didn't click End of Day)
    
    This function:
    1. Closes any active sessions
    2. Sets check_out time
    3. Marks status as 'completed'
    4. Calculates final totals
    5. Adds an audit note
    """
    try:
        # Close any active session
        active_session = work_day.get_active_session()
        if active_session:
            # Use NAIVE datetime (no timezone) since USE_TZ=False
            # Set end_time to end of the work day date
            end_of_day = datetime.combine(work_day.date, time(23, 59, 59))
            active_session.end_time = end_of_day
            active_session.notes = (active_session.notes or "") + " [Auto-closed by system]"
            active_session.save()
        
        # Set check_out if not set
        if not work_day.check_out:
            # Use the last session's end time or end of day
            last_session = work_day.sessions.order_by('-start_time').first()
            if last_session and last_session.end_time:
                work_day.check_out = last_session.end_time
            else:
                # Use NAIVE datetime
                work_day.check_out = datetime.combine(work_day.date, time(23, 59, 59))
        
        # Mark as completed
        work_day.status = 'completed'
        
        # Calculate final totals
        calculate_daily_totals_manual(work_day)
        
        # Add audit note
        now_naive = datetime.now()  # ⬅️ NAIVE datetime
        work_day.notes = (work_day.notes or "") + f"\n[Auto-closed by system on {now_naive}. Employee forgot to end work day.]"
        work_day.save()
        
        print(f"✅ Auto-closed forgotten work day: {work_day}")
        
    except Exception as e:
        print(f"❌ Error auto-closing work day {work_day.id}: {e}")
        import traceback
        traceback.print_exc()


def calculate_daily_totals_manual(work_day):
    """
    Manual calculation of daily totals to ensure it works
    Includes ACTIVE sessions in calculation
    
    IMPORTANT: We track work, break, and lunch separately.
    - total_work_time = ONLY work sessions
    - total_break_time = ONLY break sessions  
    - total_lunch_time = ONLY lunch sessions
    - productive_hours = total_work_time (already excludes breaks/lunch)
    """
    # Get ALL sessions (including active ones)
    all_sessions = work_day.sessions.all()
    
    total_work = timedelta(0)
    total_break = timedelta(0)
    total_lunch = timedelta(0)
    break_count = 0
    
    now = timezone.now()
    
    for session in all_sessions:
        # For active sessions (no end_time), calculate duration up to now
        if session.end_time:
            end_time = session.end_time
        else:
            # Active session - calculate duration up to current moment
            end_time = now
        
        if session.start_time:
            duration = end_time - session.start_time
            
            # Each session type is tracked separately
            if session.session_type == 'work':
                total_work += duration
            elif session.session_type == 'break':
                total_break += duration
                break_count += 1
            elif session.session_type == 'lunch':
                total_lunch += duration
    
    # Update work_day fields
    work_day.total_work_time = total_work
    work_day.total_break_time = total_break
    work_day.total_lunch_time = total_lunch
    work_day.break_count = break_count
    
    # productive_hours = ONLY work time (breaks/lunch already excluded)
    if total_work:
        productive_hours = total_work.total_seconds() / 3600
        work_day.productive_hours = round(productive_hours, 2)
    else:
        work_day.productive_hours = 0.0
    
    work_day.save()


def calculate_daily_stats(work_day, employee):
    """
    Calculate all daily statistics for the dashboard
    
    IMPORTANT: 
    - total_work_time already EXCLUDES breaks/lunch (it's ONLY work sessions)
    - We DON'T need to subtract breaks/lunch again
    - Payable time = total_work_time (no subtraction needed!)
    """
    # Calculate night hours manually
    night_hours = calculate_night_hours_manual(work_day)
    
    # 🔥 SAFE ACCESS: Check if fields exist and have values
    total_work_seconds = work_day.total_work_time.total_seconds() if work_day.total_work_time else 0
    total_break_seconds = work_day.total_break_time.total_seconds() if work_day.total_break_time else 0
    total_lunch_seconds = work_day.total_lunch_time.total_seconds() if work_day.total_lunch_time else 0
    
    total_work_minutes = int(total_work_seconds / 60)
    total_break_minutes = int(total_break_seconds / 60)
    total_lunch_minutes = int(total_lunch_seconds / 60)
    
    # 💰 PAYABLE TIME = total_work_time (breaks/lunch already NOT included!)
    payable_minutes = total_work_minutes
    payable_hours_decimal = decimal.Decimal(str(payable_minutes / 60))
    
    # 📅 WEEKLY HOURS = Sum of PAYABLE hours for the entire week
    weekly_hours = calculate_weekly_payable_hours(work_day)
    
    # Calculate pay breakdown based on PAYABLE TIME
    pay_breakdown = calculate_pay_breakdown_manual(
        work_day, 
        employee, 
        weekly_hours, 
        night_hours, 
        payable_hours_decimal
    )
    
    # Get hourly rate
    hourly_rate = get_hourly_rate_manual(employee)
    
    # Format durations for display
    total_work_time = f"{total_work_minutes // 60}h {total_work_minutes % 60:02d}m"
    total_break_time = f"{total_break_minutes}m"
    if total_break_minutes >= 60:
        total_break_time = f"{total_break_minutes // 60}h {total_break_minutes % 60:02d}m"
    
    total_lunch_time = f"{total_lunch_minutes}m"
    if total_lunch_minutes >= 60:
        total_lunch_time = f"{total_lunch_minutes // 60}h {total_lunch_minutes % 60:02d}m"
    
    payable_time = f"{payable_minutes // 60}h {payable_minutes % 60:02d}m"
    productive_hours = float(work_day.productive_hours or 0.0)
    
    stats = {
        'total': total_work_time,
        'payable': payable_time,
        'break': total_break_time,
        'lunch': total_lunch_time,
        'weekly_hours': weekly_hours,
        'regular_hours': pay_breakdown['regular_hours'],
        'overtime_135_hours': pay_breakdown['overtime_135_hours'],
        'overtime_200_hours': pay_breakdown['overtime_200_hours'],
        'night_hours': night_hours,
        'hourly_rate': float(hourly_rate),
        'regular_pay': f"${pay_breakdown['regular_pay']:.2f}",
        'overtime_135_pay': f"${pay_breakdown['overtime_135_pay']:.2f}",
        'overtime_200_pay': f"${pay_breakdown['overtime_200_pay']:.2f}",
        'night_pay': f"${pay_breakdown['night_pay']:.2f}",
        'total_pay': f"${pay_breakdown['total_pay']:.2f}",
        'productive_hours': productive_hours,
        'payable_hours': float(payable_hours_decimal),
        'break_count': work_day.break_count,
        'work_day_status': work_day.get_day_status() if hasattr(work_day, 'get_day_status') else 'Regular hours',
        'work_minutes': total_work_minutes,
        'break_minutes': total_break_minutes,
        'lunch_minutes': total_lunch_minutes,
        'payable_minutes': payable_minutes,
    }
    
    return stats


def calculate_weekly_payable_hours(work_day):
    """Calculate PAYABLE hours for the entire week"""
    try:
        date = work_day.date
        start_of_week = date - timedelta(days=date.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        
        weekly_workdays = WorkDay.objects.filter(
            employee=work_day.employee,
            date__range=[start_of_week, end_of_week]
        ).exclude(status__in=['absent', 'leave'])
        
        total_payable_hours = 0.0
        
        for wd in weekly_workdays:
            if wd.id == work_day.id:
                calculate_daily_totals_manual(wd)
                wd.refresh_from_db()
            
            work_seconds = wd.total_work_time.total_seconds() if wd.total_work_time else 0
            payable_hours = work_seconds / 3600
            total_payable_hours += payable_hours
        
        return round(total_payable_hours, 2)
        
    except Exception as e:
        print(f"Error calculating weekly payable hours: {e}")
        work_seconds = work_day.total_work_time.total_seconds() if work_day.total_work_time else 0
        return round(work_seconds / 3600, 2)


def calculate_night_hours_manual(work_day):
    """Calculate night hours manually (9 PM - 7 AM)"""
    try:
        night_minutes = 0
        sessions = work_day.sessions.filter(session_type='work')
        now = timezone.now()
        
        for session in sessions:
            current_time = session.start_time
            end_time = session.end_time if session.end_time else now
            
            while current_time < end_time:
                if current_time.hour >= 21 or current_time.hour < 7:
                    night_minutes += 1
                current_time += timedelta(minutes=1)
        
        return round(night_minutes / 60, 2)
    except Exception as e:
        print(f"Error calculating night hours: {e}")
        return 0.0


def get_hourly_rate_manual(employee):
    """Get hourly rate manually checking all possibilities"""
    if (hasattr(employee, 'fixed_rate') and employee.fixed_rate and 
        hasattr(employee, 'custom_base_salary') and employee.custom_base_salary):
        custom_salary = decimal.Decimal(str(employee.custom_base_salary))
        daily_rate = custom_salary / decimal.Decimal('30')
        return daily_rate / decimal.Decimal('8')
    
    if (hasattr(employee, 'position') and employee.position and 
        hasattr(employee.position, 'hour_rate') and employee.position.hour_rate):
        return decimal.Decimal(str(employee.position.hour_rate))
    
    if (hasattr(employee, 'current_campaign') and employee.current_campaign and 
        hasattr(employee.current_campaign, 'hour_rate') and employee.current_campaign.hour_rate):
        return decimal.Decimal(str(employee.current_campaign.hour_rate))
    
    return decimal.Decimal('0.00')


def calculate_pay_breakdown_manual(work_day, employee, weekly_hours, night_hours, payable_hours_decimal=None):
    """Calculate pay breakdown according to Dominican law"""
    try:
        hourly_rate = get_hourly_rate_manual(employee)
        
        if payable_hours_decimal is not None:
            daily_hours = payable_hours_decimal
        else:
            work_seconds = work_day.total_work_time.total_seconds() if work_day.total_work_time else 0
            daily_hours = decimal.Decimal(str(work_seconds / 3600))
        
        weekly_hours_dec = decimal.Decimal(str(weekly_hours))
        night_hours_dec = decimal.Decimal(str(night_hours))
        
        regular_hours = decimal.Decimal('0')
        overtime_135_hours = decimal.Decimal('0')
        overtime_200_hours = decimal.Decimal('0')
        
        if weekly_hours_dec <= decimal.Decimal('44'):
            regular_hours = daily_hours
        elif weekly_hours_dec <= decimal.Decimal('68'):
            hours_before_today = weekly_hours_dec - daily_hours
            if hours_before_today >= decimal.Decimal('44'):
                overtime_135_hours = daily_hours
            else:
                regular_hours = decimal.Decimal('44') - hours_before_today
                overtime_135_hours = daily_hours - regular_hours
        else:
            hours_before_today = weekly_hours_dec - daily_hours
            if hours_before_today >= decimal.Decimal('68'):
                overtime_200_hours = daily_hours
            elif hours_before_today >= decimal.Decimal('44'):
                overtime_135_hours = decimal.Decimal('68') - hours_before_today
                overtime_200_hours = daily_hours - overtime_135_hours
            else:
                regular_hours = decimal.Decimal('44') - hours_before_today
                remaining = daily_hours - regular_hours
                if remaining <= (decimal.Decimal('68') - decimal.Decimal('44')):
                    overtime_135_hours = remaining
                else:
                    overtime_135_hours = decimal.Decimal('68') - decimal.Decimal('44')
                    overtime_200_hours = remaining - overtime_135_hours
        
        payable_night_hours = min(night_hours_dec, daily_hours)
        
        overtime_rate_135 = hourly_rate * decimal.Decimal('1.35')
        overtime_rate_200 = hourly_rate * decimal.Decimal('2.00')
        night_rate = hourly_rate * decimal.Decimal('1.15')
        
        regular_pay = regular_hours * hourly_rate
        overtime_135_pay = overtime_135_hours * overtime_rate_135
        overtime_200_pay = overtime_200_hours * overtime_rate_200
        night_pay = payable_night_hours * night_rate
        total_pay = regular_pay + overtime_135_pay + overtime_200_pay + night_pay
        
        return {
            'regular_hours': float(regular_hours),
            'overtime_135_hours': float(overtime_135_hours),
            'overtime_200_hours': float(overtime_200_hours),
            'night_hours': float(payable_night_hours),
            'regular_pay': float(regular_pay),
            'overtime_135_pay': float(overtime_135_pay),
            'overtime_200_pay': float(overtime_200_pay),
            'night_pay': float(night_pay),
            'total_pay': float(total_pay),
            'payable_hours': float(daily_hours),
        }
    except Exception as e:
        print(f"Error calculating pay breakdown: {e}")
        return {
            'regular_hours': 0.0,
            'overtime_135_hours': 0.0,
            'overtime_200_hours': 0.0,
            'night_hours': 0.0,
            'regular_pay': 0.0,
            'overtime_135_pay': 0.0,
            'overtime_200_pay': 0.0,
            'night_pay': 0.0,
            'total_pay': 0.0,
            'payable_hours': 0.0,
        }


@require_http_methods(["POST"])
@login_required
def start_activity(request):
    """Handle status changes and update the dashboard"""
    employee = get_object_or_404(Employee, user=request.user)
    session_type = request.POST.get("session_type", "work")
    notes = request.POST.get("notes", "")

    try:
        with transaction.atomic():
            work_day = get_or_create_active_work_day(employee)
            work_day = WorkDay.objects.select_for_update().get(pk=work_day.pk)

            if session_type == "end_of_day":
                return end_work_day(request)

            active_session = work_day.get_active_session()

            if active_session and active_session.session_type == session_type:
                messages.info(request, f"You are already in {active_session.get_session_type_display()} status.")
                return agent_dashboard(request)

            if active_session:
                active_session.end_time = timezone.now()
                active_session.save()

            session = _start_new_session(work_day=work_day, session_type=session_type, notes=notes)
            calculate_daily_totals_manual(work_day)

            messages.success(request, f"Status changed to {session.get_session_type_display()}")
            return agent_dashboard(request)

    except Exception as e:
        print(f"Error in start_activity: {e}")
        import traceback
        traceback.print_exc()
        messages.error(request, f"System error: {str(e)}")
        return agent_dashboard(request)


def _start_new_session(work_day, session_type, notes=""):
    """Start a new ActivitySession"""
    try:
        session = work_day.start_work_session(session_type=session_type, notes=notes)
    except Exception:
        session = ActivitySession.objects.create(
            work_day=work_day,
            session_type=session_type,
            start_time=timezone.now(),
            notes=notes
        )
        if session_type == "work" and not work_day.check_in:
            work_day.check_in = timezone.now()
            work_day.save(update_fields=["check_in"])
    return session


@require_http_methods(["POST"])
@login_required
def end_work_day(request):
    """End complete work day - This is the ONLY way a work day should be completed"""
    employee = get_object_or_404(Employee, user=request.user)
    
    try:
        work_day = get_or_create_active_work_day(employee)
        
        active_session = work_day.get_active_session()
        if active_session:
            active_session.end_time = timezone.now()
            active_session.save()
        
        work_day.check_out = timezone.now()
        work_day.status = 'completed'
        calculate_daily_totals_manual(work_day)
        
        messages.success(request, "Work day ended successfully")
        return agent_dashboard(request)
        
    except Exception as e:
        print(f"Error ending work day: {e}")
        import traceback
        traceback.print_exc()
        messages.error(request, f"Error ending work day: {str(e)}")
        return agent_dashboard(request)

@login_required
def attendance_history(request):
    """
    Vista principal del historial de asistencia - CORREGIDA
    """
    employee = get_object_or_404(Employee, user=request.user)
    
    # Parámetros de filtrado
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    work_days = WorkDay.objects.filter(employee=employee).order_by('-date')
    
    # Filtrar por fechas
    if date_from:
        work_days = work_days.filter(date__gte=datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        work_days = work_days.filter(date__lte=datetime.strptime(date_to, '%Y-%m-%d'))

    # Paginación
    paginator = Paginator(work_days, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)


    # Preparar datos para cada día
    days_data = []
    for work_day in page_obj:
        sessions = work_day.sessions.all().order_by('start_time')
        days_data.append({
            'work_day': work_day,
            'sessions': sessions,
            'work_time': format_duration_simple(work_day.total_work_time),
            'break_time': format_duration_simple(work_day.total_break_time),
            'lunch_time': format_duration_simple(work_day.total_lunch_time),
        })
    
    context = {
        'employee': employee,
        'days_data': days_data,
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
        'total_days': work_days.count(),
    }
    
    return render(request, 'attendance/attendance_history.html', context)


@login_required
def export_attendance_excel(request):
    """
    Exporta el historial de asistencia filtrado a un archivo Excel (.xlsx)
    """
    employee = get_object_or_404(Employee, user=request.user)

    # Filtros (mismos que en attendance_history)
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    work_days = WorkDay.objects.filter(employee=employee).order_by('-date')
    if date_from:
        work_days = work_days.filter(date__gte=date_from)
    if date_to:
        work_days = work_days.filter(date__lte=date_to)

    # Crear workbook y hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance History"

    # Título
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"Attendance History - {employee.user.username}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados
    headers = ['Date', 'Work Time', 'Break Time', 'Lunch Time', 'Total Sessions']
    ws.append(headers)

    # Estilos para encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_num)].width = 18


    # Agregar datos
    for work_day in work_days:
        sessions = work_day.sessions.all().order_by('start_time')
        ws.append([
            work_day.date.strftime("%Y-%m-%d"),
            format_duration_simple(work_day.total_work_time),
            format_duration_simple(work_day.total_break_time),
            format_duration_simple(work_day.total_lunch_time),
            sessions.count(),
        ])

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attendance_history_{employee.user.username}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response



@login_required
def day_detail(request, date_str):
    """
    Vista detallada de un día específico - CORREGIDA
    """
    employee = get_object_or_404(Employee, user=request.user)
    
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        work_day = WorkDay.objects.get(employee=employee, date=date_obj)
        
        sessions = work_day.sessions.all().order_by('start_time')
        
        # Calculate all daily statistics using the same functions as dashboard
        daily_stats = calculate_daily_stats(work_day, employee)
        
        # Calculate payable hours for this specific day
        total_work_minutes = int(work_day.total_work_time.total_seconds() / 60) if work_day.total_work_time else 0
        total_break_minutes = int(work_day.total_break_time.total_seconds() / 60) if work_day.total_break_time else 0
        total_lunch_minutes = int(work_day.total_lunch_time.total_seconds() / 60) if work_day.total_lunch_time else 0
        payable_minutes = max(0, total_work_minutes - total_break_minutes - total_lunch_minutes)
        payable_hours_decimal = decimal.Decimal(str(payable_minutes / 60))
        
        # Calculate detailed pay breakdown for this specific day using PAYABLE HOURS
        pay_breakdown = calculate_pay_breakdown_manual(work_day, employee, 
                                                     daily_stats['weekly_hours'], 
                                                     daily_stats['night_hours'],
                                                     payable_hours_decimal)
        
        # Calculate daily hours breakdown for this specific day
        daily_hours_breakdown = {
            'regular_hours': pay_breakdown['regular_hours'],
            'overtime_135_hours': pay_breakdown['overtime_135_hours'],
            'overtime_200_hours': pay_breakdown['overtime_200_hours'],
            'night_hours': pay_breakdown['night_hours'],  # Use payable night hours
            'total_hours': float(work_day.productive_hours or 0.0),
            'payable_hours': pay_breakdown['payable_hours'],
        }
        
        # Get campaign info if available
        campaign = employee.current_campaign
        
        context = {
            'employee': employee,
            'work_day': work_day,
            'sessions': sessions,
            'daily_stats': daily_stats,
            'pay_breakdown': pay_breakdown,
            'daily_hours_breakdown': daily_hours_breakdown,
            'date': date_obj,
            'campaign': campaign,
            'today': date_obj,
        }
        
        return render(request, 'attendance/day_detail.html', context)
        
    except (ValueError, WorkDay.DoesNotExist):
        messages.error(request, "Day not found or no data available.")
        return redirect('attendance_history')


@login_required
def employee_profile(request, employee_id=None):
    """
    Vista del perfil del empleado
    """
    # Si no se proporciona employee_id, mostrar el perfil del usuario actual
    if employee_id is None:
        employee = get_object_or_404(Employee, user=request.user)
    else:
        employee = get_object_or_404(Employee, id=employee_id)
    
    # Verificar permisos: solo el propio empleado o superusers pueden ver información sensible
    can_view_sensitive = (request.user == employee.user) or request.user.is_superuser
    
    # Calcular estadísticas de pago si tiene permisos
    payment_stats = calculate_payment_stats(employee) if can_view_sensitive else None
    
    # Obtener campaña actual
    current_campaign = employee.current_campaign
    
    # Preparar skills como lista
    skills_list = []
    if employee.skills:
        skills_list = [skill.strip() for skill in employee.skills.split(',')]
    
    context = {
        'employee': employee,
        'current_campaign': current_campaign,
        'payment_stats': payment_stats,
        'skills_list': skills_list,
        'can_view_sensitive': can_view_sensitive,
    }
    
    return render(request, 'core/employee_profile.html', context)


def calculate_payment_stats(employee):
    """
    Calcular estadísticas de pago para el empleado
    """
    pay_method = None

    
    # Usar hourly rate del puesto como base
    if employee.current_campaign.hour_rate:
        base_rate = float(employee.current_campaign.hour_rate)
        pay_method = 'Campaign Rate / hour'

    
    print(base_rate)
    
    return {
        'pay_method': pay_method,
        'base_rate': base_rate,
        'monthly_earnings': base_rate * 160,
    }

@login_required
def edit_employee_profile(request, employee_id):
    """
    Vista para editar el perfil del empleado
    """
    employee = get_object_or_404(Employee, id=employee_id)
    
    # Verificar que el usuario puede editar este perfil
    if request.user != employee.user and not request.user.is_superuser:
        messages.error(request, "You don't have permission to edit this profile.")
        return redirect('employee_profile')
    
    if request.method == 'POST':
        form = EmployeeProfileForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated successfully!")
            return redirect('employee_profile')
    else:
        form = EmployeeProfileForm(instance=employee)
    
    context = {
        'employee': employee, 
        'form': form,
    }
    
    return render(request, 'core/edit_profile.html', context)



@login_required
def supervisor_dashboard(request):
    """
    Dashboard principal para supervisores
    """
    try:
        # Obtener el empleado que es supervisor
        supervisor = Employee.objects.get(user=request.user, is_supervisor=True)
    except Employee.DoesNotExist:
        messages.error(request, "You don't have supervisor privileges.")
        return redirect('employee_profile')
    
    # Obtener los miembros del equipo con información del usuario
    team_members = Employee.objects.filter(
        supervisor=supervisor,
        is_active=True
    ).select_related(
        'user', 'position', 'department', 'current_campaign'
    )

    # Estadísticas del equipo
    total_team_members = team_members.count()
    logged_in_count = team_members.filter(is_logged_in=True).count()
    active_in_campaign = team_members.filter(current_campaign__isnull=False).count()

    
    # Obtener WorkDays de hoy para el equipo
    today = timezone.now().date()
    team_workdays = WorkDay.objects.filter(
        employee__in=team_members,
        date=today
    ).select_related('employee')
    
    # Obtener información de sesiones (logins) más recientes
    from django.contrib.sessions.models import Session
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    
    # Preparar datos para cada miembro del equipo
    team_data = []
    for member in team_members:
        try:
            workday_today = team_workdays.get(employee=member)
            current_session = workday_today.get_active_session()
            daily_stats = calculate_daily_stats(workday_today)
        except WorkDay.DoesNotExist:
            workday_today = None
            current_session = None
            daily_stats = None
        
        team_data.append({
            'employee': member,
            'workday': workday_today,
            'current_session': current_session,
            'formatted_session': workday_today.get_formatted_session() if workday_today else None,
            'daily_stats': daily_stats,
        })
            
    context = {
        'supervisor': supervisor,
        'team_data': team_data,
        'total_team_members': total_team_members,
        'logged_in_count': logged_in_count,
        'active_in_campaign': active_in_campaign,
        'today': today,
    }
    
    return render(request, 'supervisor/supervisor_dashboard.html', context)


from tempfile import NamedTemporaryFile

login_required
def export_team_report_excel(request):
    """
    Genera y descarga directamente el reporte de equipo en Excel (sin email).
    """
    try:
        supervisor = Employee.objects.get(user=request.user, is_supervisor=True)
    except Employee.DoesNotExist:
        messages.error(request, "You don't have supervisor privileges.")
        return redirect('employee_profile')

    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if not date_from or not date_to:
        messages.error(request, "Please select valid dates.")
        return redirect("supervisor_dashboard")

    # Convertir fechas
    date_from_dt = datetime.strptime(date_from, "%Y-%m-%d").date()
    date_to_dt = datetime.strptime(date_to, "%Y-%m-%d").date()

    # Obtener miembros del equipo
    team_members = Employee.objects.filter(supervisor=supervisor)

    work_days = (
        WorkDay.objects.filter(employee__in=team_members, date__range=(date_from_dt, date_to_dt))
        .select_related("employee", "employee__user")
        .prefetch_related("sessions")
        .order_by("date", "employee__user__last_name")
    )

    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Team Report"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    # Títulos
    ws.merge_cells("A1:Q1")
    ws["A1"] = f"Team Report - Supervisor: {supervisor.user.get_full_name()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:Q2")
    ws["A2"] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].alignment = center

    ws.append([])

    headers = [
        "Agent Name", "Date", "Attendance Status", "Week",
        "Time In",
        "Break 1-start", "Break 1-end", "Break1 Duration",
        "Break 2-start", "Break 2-end", "Break2 Duration",
        "Lunch start", "Lunch end", "Lunch Duration",
        "Time Out",
        "Total Hours",
        "Notes",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=4, column=col).font = bold
        ws.cell(row=4, column=col).alignment = center

    # Helpers
    def fmt_time(t):
        return t.strftime("%H:%M") if t else "-"

    def fmt_duration(s):
        if not s or not s.start_time or not s.end_time:
            return "-"
        delta = s.end_time - s.start_time
        minutes = delta.total_seconds() // 60
        seconds = int(delta.total_seconds() % 60)
        return f"{int(minutes):02d}:{seconds:02d}"

    def fmt_total(td):
        if td is None:
            return "0.00"
        return f"{td.total_seconds() / 3600:.2f}"

    # Filas
    for wd in work_days:
        sessions = wd.sessions.all()
        work_sessions = sessions.filter(session_type="work").order_by("start_time")
        break_sessions = sessions.filter(session_type="break").order_by("start_time")
        lunch_sessions = sessions.filter(session_type="lunch").order_by("start_time")

        time_in = work_sessions.first().start_time if work_sessions else None
        time_out = work_sessions.last().end_time if work_sessions else None
        break1 = break_sessions[0] if len(break_sessions) >= 1 else None
        break2 = break_sessions[1] if len(break_sessions) >= 2 else None
        lunch = lunch_sessions[0] if lunch_sessions else None

        ws.append([
            wd.employee.user.get_full_name(),
            wd.date.strftime("%m/%d/%Y"),
            wd.get_day_status(),
            wd.date.isocalendar()[1],
            fmt_time(time_in),
            fmt_time(break1.start_time if break1 else None),
            fmt_time(break1.end_time if break1 else None),
            fmt_duration(break1) if break1 else "-",
            fmt_time(break2.start_time if break2 else None),
            fmt_time(break2.end_time if break2 else None),
            fmt_duration(break2) if break2 else "-",
            fmt_time(lunch.start_time if lunch else None),
            fmt_time(lunch.end_time if lunch else None),
            fmt_duration(lunch) if lunch else "-",
            fmt_time(time_out),
            fmt_total(wd.total_work_time),
            wd.notes or "",
        ])

    # Ajustar ancho columnas
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2

    # Crear nombre del archivo
    filename = f"team_report_{supervisor.user.username}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Guardar en archivo temporal
    with NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        wb.save(tmp_file.name)
        temp_path = tmp_file.name

    try:
        # Leer el archivo y devolverlo como respuesta HTTP
        with open(temp_path, 'rb') as f:
            response = HttpResponse(
                f.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    finally:
        # Limpiar archivo temporal
        if os.path.exists(temp_path):
            os.remove(temp_path)

@login_required
def team_attendance_history(request):
    """
    Historial de asistencia del equipo completo
    """
    try:
        supervisor = Employee.objects.get(user=request.user, is_supervisor=True)
    except Employee.DoesNotExist:
        messages.error(request, "You don't have supervisor privileges.")
        return redirect('employee_profile')
    
    # Parámetros de filtrado
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    employee_id = request.GET.get('employee')
    
    # Obtener miembros del equipo
    team_members = Employee.objects.filter(supervisor=supervisor)
    
    # Obtener WorkDays del equipo
    work_days = WorkDay.objects.filter(employee__in=team_members).order_by('-date')
    
    # Aplicar filtros
    if date_from:
        work_days = work_days.filter(date__gte=date_from)
    if date_to:
        work_days = work_days.filter(date__lte=date_to)
    if employee_id:
        work_days = work_days.filter(employee_id=employee_id)
    
    # Paginación
    paginator = Paginator(work_days, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Preparar datos para cada día
    days_data = []
    for work_day in page_obj:
        sessions = work_day.sessions.all().order_by('start_time')
        days_data.append({
            'work_day': work_day,
            'sessions': sessions,
            'work_time': format_duration_simple(work_day.total_work_time),
            'break_time': format_duration_simple(work_day.total_break_time),
            'lunch_time': format_duration_simple(work_day.total_lunch_time),
        })
    
    context = {
        'supervisor': supervisor,
        'team_members': team_members,
        'days_data': days_data,
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
        'selected_employee': employee_id,
        'total_days': work_days.count(),
    }
    
    return render(request, 'supervisor/team_attendance_history.html', context)

@login_required
def employee_attendance_detail(request, employee_id):
    """
    Vista detallada de asistencia de un empleado específico
    Muestra lista de días o detalle de un día específico
    """
    try:
        employee = Employee.objects.get(id=employee_id, is_active=True)
        
        # Verificar permisos
        if not (request.user.is_superuser or 
                Employee.objects.filter(user=request.user, is_supervisor=True).exists()):
            messages.error(request, "You don't have permission to view this page.")
            return redirect('agent_dashboard')
            
        supervisor = Employee.objects.filter(user=request.user, is_supervisor=True).first()
        
    except Employee.DoesNotExist:
        messages.error(request, "Employee not found.")
        return redirect('supervisor_dashboard')
    
    # ============================================================================
    # DAY DETAIL VIEW - Si se solicita un día específico
    # ============================================================================
    date_str = request.GET.get('day')
    if date_str:
        try:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            work_day = WorkDay.objects.get(employee=employee, date=date_obj)
            
            # Forzar recalculo de totales
            calculate_daily_totals_manual(work_day)
            work_day.refresh_from_db()
            
            # Obtener sesiones
            sessions = work_day.sessions.all().order_by('start_time')
            
            # Calcular estadísticas del día
            daily_stats = calculate_daily_stats(work_day, employee)
            
            context = {
                'supervisor': supervisor,
                'employee': employee,
                'work_day': work_day,
                'sessions': sessions,
                'date': date_obj,
                'daily_stats': daily_stats,
                'is_day_detail': True,
            }
            
            return render(request, 'supervisor/employee_day_detail.html', context)
        
        except WorkDay.DoesNotExist:
            messages.error(request, f"No work day found for {date_obj}")
        except Exception as e:
            messages.error(request, f"Error loading day details: {str(e)}")
            import traceback
            traceback.print_exc()
    
    # ============================================================================
    # LIST VIEW - Mostrar lista de días
    # ============================================================================
    
    # Parámetros de filtrado
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Obtener WorkDays del empleado
    work_days = WorkDay.objects.filter(
        employee=employee
    ).select_related(
        'employee__position', 
        'employee__department',
        'employee__current_campaign'
    ).prefetch_related(
        'sessions'
    ).order_by('-date')
    
    # Aplicar filtros
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            work_days = work_days.filter(date__gte=date_from_obj)
        except ValueError:
            messages.warning(request, "Invalid 'from' date format")
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            work_days = work_days.filter(date__lte=date_to_obj)
        except ValueError:
            messages.warning(request, "Invalid 'to' date format")
    
    # Paginación
    paginator = Paginator(work_days, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # ============================================================================
    # Calcular estadísticas del empleado (solo para días filtrados)
    # ============================================================================
    total_work_days = work_days.count()
    
    # Calcular tiempo total de trabajo
    total_work_seconds = 0
    for wd in work_days:
        if wd.total_work_time:
            total_work_seconds += wd.total_work_time.total_seconds()
    
    # Calcular promedio
    avg_work_seconds = total_work_seconds / total_work_days if total_work_days > 0 else 0
    
    # Formatear tiempos
    total_work_time_formatted = format_duration_hours(timedelta(seconds=total_work_seconds))
    avg_work_time_formatted = format_duration_hours(timedelta(seconds=avg_work_seconds))
    
    context = {
        'supervisor': supervisor,
        'employee': employee,
        'page_obj': page_obj,
        'date_from': date_from,
        'date_to': date_to,
        'total_work_days': total_work_days,
        'total_work_time': total_work_time_formatted,
        'avg_work_time': avg_work_time_formatted,
        'is_day_detail': False,
    }
    
    return render(request, 'supervisor/employee_attendance_detail.html', context)


def format_duration_hours(duration):
    """
    Formato: "8h 30m"
    """
    if not duration:
        return "0h 0m"
    
    total_seconds = int(duration.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    
    return f"{hours}h {minutes}m"


def format_duration_simple(duration):
    """
    Alias para compatibilidad
    """
    return format_duration_hours(duration)


@login_required
def export_employee_attendance_excel(request, employee_id):
    """Exporta los registros de asistencia de un empleado específico a Excel (.xlsx) con detalle de breaks y lunch."""
    try:
        supervisor = Employee.objects.get(user=request.user, is_supervisor=True)
    except Employee.DoesNotExist:
        messages.error(request, "You don't have supervisor privileges.")
        return redirect('employee_profile')

    employee = get_object_or_404(Employee, id=employee_id, supervisor=supervisor)

    # Filtros de fechas
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if not date_from or not date_to:
        messages.error(request, "You must select a valid start and end date before generating the report.")
        return redirect(request.META.get('HTTP_REFERER', 'employee_profile'))

    try:
        date_from_parsed = datetime.strptime(date_from, "%Y-%m-%d").date()
        date_to_parsed = datetime.strptime(date_to, "%Y-%m-%d").date()
        if date_from_parsed > date_to_parsed:
            messages.error(request, "The start date cannot be later than the end date.")
            return redirect(request.META.get('HTTP_REFERER', 'employee_profile'))
    except ValueError:
        messages.error(request, "Invalid date format. Please select valid dates.")
        return redirect(request.META.get('HTTP_REFERER', 'employee_profile'))

    work_days = WorkDay.objects.filter(
        employee=employee,
        date__range=(date_from_parsed, date_to_parsed)
    ).prefetch_related('sessions').order_by('date')

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")

    # Formateadores
    def format_timedelta(td):
        if not td:
            return "00:00:00"
        total_seconds = int(td.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02}:{minutes:02}:{seconds:02}"

    def format_duration(s):
        if s.start_time and s.end_time:
            return format_timedelta(s.end_time - s.start_time)
        return "—"

    # Encabezado principal
    ws.merge_cells('A1:P1')
    ws['A1'] = f"Attendance Report for {employee.user.get_full_name()}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:P2')
    ws['A2'] = f"Date range: {date_from} to {date_to}"
    ws['A2'].alignment = center_align

    ws.merge_cells('A3:P3')
    ws['A3'] = f"Generated at: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'].alignment = center_align
    ws.append([])

    # Encabezados
    headers = [
        "Date", "Day", "Status", "Check In", "Check Out", "Total Work Time",
        "Break1 Start", "Break1 End", "Break1 Duration",
        "Break2 Start", "Break2 End", "Break2 Duration",
        "Lunch Start", "Lunch End", "Lunch Duration",
        "Work Sessions Count"
    ]
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align

    total_seconds = 0

    for wd in work_days:
        total_time = wd.total_work_time or timedelta()
        total_seconds += total_time.total_seconds()

        check_in = wd.check_in.strftime("%I:%M %p") if wd.check_in else "—"
        check_out = wd.check_out.strftime("%I:%M %p") if wd.check_out else "—"

        breaks = [s for s in wd.sessions.all() if s.session_type == 'break'][:2]
        lunch = [s for s in wd.sessions.all() if s.session_type == 'lunch'][:1]

        def fmt_time(t): return t.strftime("%I:%M %p") if t else "—"
        def fmt_duration(s): return s.end_time - s.start_time if s.start_time and s.end_time else None

        break1 = fmt_duration(breaks[0]) if len(breaks) > 0 else None
        break2 = fmt_duration(breaks[1]) if len(breaks) > 1 else None
        lunch_d = fmt_duration(lunch[0]) if lunch else None

        work_sessions_count = wd.sessions.filter(session_type='work').count()

        # Convertir total_time a formato Excel (fracción de día)
        excel_total_time = total_time.total_seconds() / 86400  # 1 día = 86400 s

        row = [
            wd.date.strftime("%Y-%m-%d"),
            wd.date.strftime("%A"),
            wd.get_status_display(),
            check_in,
            check_out,
            excel_total_time,  # numérico, no string
            fmt_time(breaks[0].start_time) if len(breaks) > 0 else "—",
            fmt_time(breaks[0].end_time) if len(breaks) > 0 else "—",
            str(break1) if break1 else "—",
            fmt_time(breaks[1].start_time) if len(breaks) > 1 else "—",
            fmt_time(breaks[1].end_time) if len(breaks) > 1 else "—",
            str(break2) if break2 else "—",
            fmt_time(lunch[0].start_time) if lunch else "—",
            fmt_time(lunch[0].end_time) if lunch else "—",
            str(lunch_d) if lunch_d else "—",
            work_sessions_count
        ]
        ws.append(row)

        # Aplica formato h:mm a la columna "Total Work Time" (columna 6)
        ws.cell(row=ws.max_row, column=6).number_format = "[h]:mm"

    # --- Fila resumen ---
    ws.append([])
    total_time_final = timedelta(seconds=total_seconds)
    excel_total_final = total_time_final.total_seconds() / 86400
    ws.append(["", "", "", "", "TOTAL WORK TIME:", excel_total_final])
    ws.cell(row=ws.max_row, column=5).font = bold_font
    ws.cell(row=ws.max_row, column=6).font = bold_font
    ws.cell(row=ws.max_row, column=6).number_format = "[h]:mm"

    # Ajustar anchos
    for i, column_cells in enumerate(ws.columns, 1):
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(i)].width = length + 2

    # Crear respuesta
    safe_name = f"{employee.user.first_name}_{employee.user.last_name}".replace(" ", "_")
    filename = f"attendance_{safe_name}_{timezone.now().strftime('%Y%m%d')}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def supervisor_day_detail(request, employee_id, date_str):
    """
    Vista detallada de un día específico
    Acceso: Cualquier supervisor o admin puede ver cualquier empleado
    """
    
    # ============================================================================
    # VERIFICAR PERMISOS - Simple: ¿Eres supervisor o admin?
    # ============================================================================
    is_supervisor = False
    user_employee = None
    
    try:
        user_employee = Employee.objects.get(user=request.user)
        is_supervisor = user_employee.is_supervisor
    except Employee.DoesNotExist:
        pass
    
    # Permitir acceso si es:
    # 1. Superuser/Admin
    # 2. Staff
    # 3. Supervisor (cualquier supervisor)
    has_access = (
        request.user.is_superuser or 
        request.user.is_staff or 
        is_supervisor or
        request.user.groups.filter(name__in=['Supervisors', 'HR', 'Managers']).exists()
    )
    
    if not has_access:
        messages.error(request, "You don't have permission to view attendance details.")
        return redirect('agent_dashboard')
    
    # ============================================================================
    # CARGAR DATOS DEL DÍA
    # ============================================================================
    
    # Obtener empleado (sin restricciones de supervisor)
    employee = get_object_or_404(Employee, id=employee_id)
    
    try:
        # Parse fecha
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Obtener WorkDay
        work_day = get_object_or_404(WorkDay, employee=employee, date=date_obj)
        
        # 🔥 IMPORTANTE: Forzar recalculo usando las funciones centralizadas
        calculate_daily_totals_manual(work_day)
        work_day.refresh_from_db()
        
        # Obtener sesiones
        sessions = work_day.sessions.all().order_by('start_time')
        
        # 📊 Calcular estadísticas (misma función que agent_dashboard)
        daily_stats = calculate_daily_stats(work_day, employee)
        
        # ============================================================================
        # Calcular tiempos adicionales para el template
        # ============================================================================
        
        # Tiempo total del día (check_in a check_out)
        total_day_duration = timedelta(0)
        if work_day.check_in:
            if work_day.check_out:
                total_day_duration = work_day.check_out - work_day.check_in
            else:
                # Si no hay check_out, calcular hasta ahora
                last_session = sessions.order_by('-start_time').first()
                if last_session:
                    end_time = last_session.end_time or timezone.now()
                    total_day_duration = end_time - work_day.check_in
                else:
                    total_day_duration = timezone.now() - work_day.check_in
        
        # Tiempos ya calculados en work_day
        total_work = work_day.total_work_time or timedelta(0)
        total_break = work_day.total_break_time or timedelta(0)
        total_lunch = work_day.total_lunch_time or timedelta(0)
        
        # Payable hours = total_work (ya excluye breaks/lunch)
        payable_hours = total_work
        
        # Timestamps
        day_start = work_day.check_in
        day_end = work_day.check_out
        
        # ============================================================================
        # Permisos para acciones en el template
        # ============================================================================
        can_edit = request.user.is_superuser or is_supervisor or request.user.is_staff
        can_delete = request.user.is_superuser
        
        context = {
            'supervisor': user_employee,
            'employee': employee,
            'work_day': work_day,
            'sessions': sessions,
            'date': date_obj,
            
            # Duraciones
            'total_day_duration': total_day_duration,
            'total_work': total_work,
            'total_break': total_break,
            'total_lunch': total_lunch,
            'payable_hours': payable_hours,
            
            # Timestamps
            'day_start': day_start,
            'day_end': day_end,
            
            # Estadísticas completas
            'daily_stats': daily_stats,
            
            # Permisos
            'can_edit': can_edit,
            'can_delete': can_delete,
        }

        return render(request, 'supervisor/supervisor_day_detail.html', context)

    except ValueError:
        messages.error(request, "Invalid date format. Use YYYY-MM-DD.")
        return redirect('employee_attendance_detail', employee_id=employee_id)
        
    except WorkDay.DoesNotExist:
        messages.error(request, f"No work day found for {employee.user.get_full_name()} on {date_str}.")
        return redirect('employee_attendance_detail', employee_id=employee_id)
        
    except Exception as e:
        messages.error(request, f"Error loading day details: {str(e)}")
        import traceback
        traceback.print_exc()
        return redirect('employee_attendance_detail', employee_id=employee_id)
    

@login_required(login_url='/accounts/login/')
def edit_session(request, pk):
    session = get_object_or_404(ActivitySession, pk=pk)


    # if not is_supervisor(request.user):
    #     return HttpResponseForbidden(render(request, "403.html"))
    
    if request.method == 'POST':
        form = ActivitySessionForm(request.POST, instance=session)
        if form.is_valid():
            form.save()
            return redirect('supervisor_day_detail', employee_id=session.work_day.employee.id, date_str=session.work_day.date.strftime('%Y-%m-%d'))
    else:
        form = ActivitySessionForm(instance=session)
    return render(request, 'supervisor/edit_session.html', {'form': form, 'session': session})


@login_required(login_url='/accounts/login/')
def delete_session(request, pk):
    session = get_object_or_404(ActivitySession, pk=pk)

    # if not is_supervisor(request.user):
    #     return HttpResponseForbidden(render(request, "403.html"))
    
    if request.method == 'POST':
        # Guardamos la información de la sesión antes de eliminarla para la redirección
        work_day = session.work_day
        employee_id = work_day.employee.id
        date_str = work_day.date.strftime('%Y-%m-%d')
        
        # Eliminar la sesión
        session.delete()
        
        # Recalcular los totales del día
        work_day.calculate_daily_totals()
        
        # Actualizar información de ajustes
        work_day.update_adjustment_info(request.user)
        
        return redirect('supervisor_day_detail', employee_id=employee_id, date_str=date_str)
    
    # Para solicitudes GET, mostramos la página de confirmación
    return render(request, 'supervisor/delete_session.html', {'session': session})

@login_required(login_url='/accounts/login/')
def workday_editor_view(request, workday_id):
    workday = get_object_or_404(WorkDay, id=workday_id)
    sessions = workday.sessions.all()

    sessions_data = []
    for s in sessions:
        # Manejar end_time que puede ser None (sesión en curso)
        end_time_formatted = s.end_time.strftime('%H:%M') if s.end_time else "En curso"
        
        sessions_data.append({
            'id': s.id,
            'type': s.get_session_type_display(),
            'type_code': s.session_type,
            'start': s.start_time.strftime('%H:%M'),
            'end': end_time_formatted,
        })

    referer = request.META.get('HTTP_REFERER', '/')

    return render(request, 'supervisor/workday_timeline.html', {
        'workday': workday,
        'sessions': sessions,
        'sessions_json': json.dumps(sessions_data),
        'referer': referer
    })


@login_required(login_url='/accounts/login/')
@require_http_methods(["POST"])
def update_session(request, session_id):
    session = get_object_or_404(ActivitySession, id=session_id)
    
    # if not is_supervisor(request.user):
    #     return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    
    try:
        new_start = request.POST.get('start')
        new_end = request.POST.get('end')
        adjustment_reason = request.POST.get('adjustment_reason', '')
        
        # Convertir tiempos
        workday_date = session.work_day.date
        new_start_time = datetime.combine(workday_date, datetime.strptime(new_start, '%H:%M').time())
        new_end_time = datetime.combine(workday_date, datetime.strptime(new_end, '%H:%M').time())
        
        # Ajustar la sesión
        session.adjust_times(
            new_start_time=new_start_time,
            new_end_time=new_end_time,
            adjusted_by=request.user,
            notes=adjustment_reason
        )
        
        # Registrar en el WorkDay
        session.work_day.add_adjustment_record(
            adjusted_by=request.user,
            reason=adjustment_reason,
            sessions_affected=[session.id]
        )
        
        # Recalcular y preparar respuesta
        workday = session.work_day
        workday.calculate_daily_totals()
        
        # Preparar datos de respuesta
        updated_sessions = []
        for s in workday.sessions.all().order_by('start_time'):
            updated_sessions.append({
                'id': s.id,
                'start': s.start_time.strftime('%H:%M'),
                'end': s.end_time.strftime('%H:%M'),
                'duration': s.duration_minutes,
                'is_adjusted': s.is_adjusted
            })
        
        return JsonResponse({
            'success': True,
            'updated_sessions': updated_sessions,
            'total_work': workday.total_work_minutes,
            'total_breaks': workday.total_break_minutes,
            'total_lunch': workday.total_lunch_minutes,
            'adjustment_reason': adjustment_reason
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})
    


#Occurrence crud
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models.functions import Concat
@login_required
def occurrence_list(request):
    """Lista todas las ocurrencias con paginación y filtros"""
    
    # Get the employee for the current user
    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('dashboard')
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    campaign_filter = request.GET.get('campaign', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    

    if employee.is_supervisor or employee.is_it or request.user.is_superuser:
        occurrence_list = Occurrence.objects.all()
    else:
        occurrence_list = Occurrence.objects.filter(employee=employee)
    
    # Apply date filters
    if date_from:
        occurrence_list = occurrence_list.filter(date__gte=date_from)  # Use date field
    
    if date_to:
        occurrence_list = occurrence_list.filter(date__lte=date_to)  # Use date field
    
    # Continue with other filters...
    occurrence_list = occurrence_list.order_by('-date', '-start_time')  # Order by date first
    
    # Apply search filter
    if search_query:
        if employee.is_supervisor or employee.is_it or request.user.is_superuser:
            occurrence_list = occurrence_list.annotate(
                full_name_db=Concat(
                    'employee__user__first_name',
                    Value(' '),
                    'employee__user__last_name'
                )
            ).filter(
                Q(full_name_db__icontains=search_query) |
                Q(employee__user__first_name__icontains=search_query) |
                Q(employee__user__last_name__icontains=search_query)
            )
        else:
            occurrence_list = occurrence_list.filter(
                Q(comment__icontains=search_query) |
                Q(occurrence_type__icontains=search_query)
            )
    
    # Apply campaign filter
    if campaign_filter:
        if employee.is_supervisor or employee.is_it or request.user.is_superuser:
            occurrence_list = occurrence_list.filter(
                employee__current_campaign__id=campaign_filter
            )
    
    # Build query parameters for pagination links
    query_params = ''
    if search_query:
        query_params += f'&search={search_query}'
    if campaign_filter:
        query_params += f'&campaign={campaign_filter}'
    if date_from:
        query_params += f'&date_from={date_from}'
    if date_to:
        query_params += f'&date_to={date_to}'
    
    # Pagination
    paginator = Paginator(occurrence_list, 20)  # Show 20 occurrences per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get campaigns for filter dropdown (only for supervisors)
    if employee.is_supervisor or employee.is_it or request.user.is_superuser:
        campaigns = Campaign.objects.all()
    else:
        campaigns = None
    
    context = {
        'occurrences': page_obj,  # Changed from 'page_obj' to 'occurrences'
        'page_obj': page_obj,  # Keep both for compatibility
        'search_query': search_query,
        'campaign_filter': campaign_filter,
        'date_from': date_from,
        'date_to': date_to,
        'campaigns': campaigns,
        'is_supervisor': employee.is_supervisor or employee.is_it or request.user.is_superuser,
        'query_params': query_params,  # Added this
    }
    
    return render(request, 'occurrences/occurrence_list.html', context)

@login_required
def occurrence_create(request):
    employee = Employee.objects.get(user=request.user)

    if request.method == "POST":
        session_occurrence = request.POST.get("session_occurrence")
        start_time_str = request.POST.get("start_time")
        end_time_str = request.POST.get("end_time")
        comment = request.POST.get("comment")

        # Collect validation errors
        errors = []
        
        if not session_occurrence:
            errors.append('You must select an Occurrence')
        
        if not start_time_str or not end_time_str:
            errors.append('Start time and end time are required.')
        
        # if not comment:
        #     errors.append('Comment is required.')

        # If there are any validation errors, show them and redirect
        if errors:
            for error in errors:
                messages.info(request, error)
            return redirect("agent_dashboard")

        try:
            # Only parse time if strings are not empty
            start_time = datetime.strptime(start_time_str, "%H:%M").time()
            end_time = datetime.strptime(end_time_str, "%H:%M").time()
        except ValueError:
            messages.info(request, "Invalid time format. Please use HH:MM format (e.g., 09:30, 14:45)")
            return redirect("agent_dashboard")

        # Create the occurrence
        Occurrence.objects.create(
            employee=employee,
            occurrence_type=session_occurrence,
            start_time=start_time,
            end_time=end_time,
            comment=comment,
            created_at=timezone.now(),
            date=timezone.now()
        )

        messages.success(request, "Occurrence registered successfully!")
    
    return redirect("agent_dashboard")


@login_required
def occurrence_update(request, pk):

    occurrence = get_object_or_404(Occurrence, pk=pk)
    
    if request.method == 'POST':
        form = OccurrenceForm(request.POST, instance=occurrence)
        if form.is_valid():
            form.save()
            messages.success(request, 'Occurrence updated successfully!')
            return redirect('occurrence_list')
    else:
        form = OccurrenceForm(instance=occurrence)
    
    return render(request, 'occurrences/occurrence_form.html', {
        'form': form,
        'title': 'Update Occurrence',
        'occurrence': occurrence
    })

@login_required
def occurrence_detail(request, pk):

    occurrence = get_object_or_404(Occurrence, pk=pk, employee=request.user)
    return render(request, 'occurrences/occurrence_detail.html', {
        'occurrence': occurrence
    })

@login_required
def occurrence_delete(request, occurrence_id):
    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        messages.error(request, "Employee profile not found.")
        return redirect('dashboard')
    
    # Get the occurrence for this specific employee
    occurrence = get_object_or_404(Occurrence, id=occurrence_id, employee=employee)
    
    if request.method == 'POST':
        occurrence.delete()
        messages.success(request, 'Occurrence deleted successfully!')
        return redirect('occurrence_list')
    
    # For GET request, show confirmation page
    context = {
        'occurrence': occurrence
    }

    return HttpResponse()

@login_required
@user_passes_test(is_supervisor)
def bulk_create_workday(request):
    """
    View to create work days for multiple employees
    Handles system issues, forgotten check-ins, etc.
    """
    if request.method == 'POST':
        return handle_bulk_workday_creation(request)
    
    # GET request - show form
    from attendance.models import Campaign
    
    # Get all campaigns for filtering
    campaigns = Campaign.objects.filter(is_active=True)
    
    context = {
        'campaigns': campaigns,
        'today': timezone.now().date(),
    }
    
    return render(request, 'attendance/bulk_create_workday.html', context)


@login_required
@user_passes_test(is_supervisor)
def load_employees_dropdown(request):
    """
    HTMX endpoint to load employees based on campaign filter
    """
    from attendance.models import Employee
    
    campaign_id = request.GET.get('campaign_id')
    
    employees = Employee.objects.filter(
        is_active=True
    ).select_related('position', 'current_campaign')
    
    if campaign_id:
        employees = employees.filter(current_campaign_id=campaign_id)
    
    return render(request, 'attendance/partials/employees_dropdown.html', {
        'employees': employees
    })


@login_required
@user_passes_test(is_supervisor)
def add_break_session(request):
    """
    HTMX endpoint to add a break session row
    """
    index = request.GET.get('index', 0)
    return render(request, 'attendance/partials/break_session.html', {
        'index': index
    })


@login_required
@user_passes_test(is_supervisor)
def add_lunch_session(request):
    """
    HTMX endpoint to add a lunch session row
    """
    index = request.GET.get('index', 0)
    return render(request, 'attendance/partials/lunch_session.html', {
        'index': index
    })


def handle_bulk_workday_creation(request):
    """
    Process the bulk work day creation form
    """
    try:
        # Get form data
        selected_employees = request.POST.getlist('employees')
        work_date = request.POST.get('work_date')
        
        # Time inputs
        check_in_time = request.POST.get('check_in_time')
        check_out_time = request.POST.get('check_out_time')
        
        # Break sessions
        break_sessions = []
        i = 0
        while True:
            break_start = request.POST.get(f'break_{i}_start')
            break_end = request.POST.get(f'break_{i}_end')
            if not break_start or not break_end:
                break
            break_sessions.append({
                'start': break_start,
                'end': break_end,
                'type': 'break'
            })
            i += 1
        
        # Lunch sessions
        lunch_sessions = []
        i = 0
        while True:
            lunch_start = request.POST.get(f'lunch_{i}_start')
            lunch_end = request.POST.get(f'lunch_{i}_end')
            if not lunch_start or not lunch_end:
                break
            lunch_sessions.append({
                'start': lunch_start,
                'end': lunch_end,
                'type': 'lunch'
            })
            i += 1
        
        reason = request.POST.get('reason', 'Manual creation by supervisor')
        
        # Validate inputs
        if not selected_employees:
            messages.error(request, "Please select at least one employee")
            return redirect('bulk_create_workday')
        
        if not work_date or not check_in_time or not check_out_time:
            messages.error(request, "Date, check-in, and check-out times are required")
            return redirect('bulk_create_workday')
        
        # Convert date string to date object
        work_date_obj = datetime.strptime(work_date, '%Y-%m-%d').date()
        
        # Parse times
        check_in_datetime = parse_datetime(work_date_obj, check_in_time)
        check_out_datetime = parse_datetime(work_date_obj, check_out_time)
        
        # If check_out is earlier than check_in, it means next day
        if check_out_datetime <= check_in_datetime:
            check_out_datetime += timedelta(days=1)
        
        # Create work days in transaction
        created_count = 0
        errors = []
        
        with transaction.atomic():
            for employee_id in selected_employees:
                try:
                    from attendance.models import Employee, WorkDay
                    employee = Employee.objects.get(id=employee_id)
                    
                    # Check if work day already exists
                    existing = WorkDay.objects.filter(
                        employee=employee,
                        date=work_date_obj
                    ).first()
                    
                    if existing:
                        errors.append(f"{employee.get_full_name()}: Work day already exists")
                        continue
                    
                    # Create work day
                    work_day = create_workday_with_sessions(
                        employee=employee,
                        work_date=work_date_obj,
                        check_in=check_in_datetime,
                        check_out=check_out_datetime,
                        break_sessions=break_sessions,
                        lunch_sessions=lunch_sessions,
                        reason=reason,
                        created_by=request.user
                    )
                    
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"{employee.get_full_name()}: {str(e)}")
        
        # Show results
        if created_count > 0:
            messages.success(
                request, 
                f"Successfully created {created_count} work day(s)"
            )
        
        if errors:
            for error in errors:
                messages.warning(request, error)
        
        return redirect('bulk_create_workday')
        
    except Exception as e:
        messages.error(request, f"Error creating work days: {str(e)}")
        return redirect('bulk_create_workday')


def create_workday_with_sessions(employee, work_date, check_in, check_out, 
                                  break_sessions, lunch_sessions, reason, created_by):
    """
    Create a complete work day with all sessions
    """
    from attendance.models import WorkDay, ActivitySession
    
    # Create WorkDay
    work_day = WorkDay.objects.create(
        employee=employee,
        date=work_date,
        check_in=check_in,
        check_out=check_out,
        status='completed',
        notes=f"Created manually: {reason}",
    )
    
    # Build timeline of all sessions
    all_sessions = []
    
    # Add break sessions
    for break_session in break_sessions:
        break_start = parse_datetime(work_date, break_session['start'])
        break_end = parse_datetime(work_date, break_session['end'])
        if break_end <= break_start:
            break_end += timedelta(days=1)
        
        all_sessions.append({
            'start': break_start,
            'end': break_end,
            'type': 'break'
        })
    
    # Add lunch sessions
    for lunch_session in lunch_sessions:
        lunch_start = parse_datetime(work_date, lunch_session['start'])
        lunch_end = parse_datetime(work_date, lunch_session['end'])
        if lunch_end <= lunch_start:
            lunch_end += timedelta(days=1)
        
        all_sessions.append({
            'start': lunch_start,
            'end': lunch_end,
            'type': 'lunch'
        })
    
    # Sort all non-work sessions by start time
    all_sessions.sort(key=lambda x: x['start'])
    
    # Create work sessions (gaps between check_in, breaks/lunch, and check_out)
    current_time = check_in
    
    for session in all_sessions:
        # If there's a gap before this break/lunch, create work session
        if current_time < session['start']:
            ActivitySession.objects.create(
                work_day=work_day,
                session_type='work',
                start_time=current_time,
                end_time=session['start'],
                notes=f"Auto-created by {created_by.username}",
                auto_created=True
            )
        
        # Create break/lunch session
        ActivitySession.objects.create(
            work_day=work_day,
            session_type=session['type'],
            start_time=session['start'],
            end_time=session['end'],
            notes=f"Manual entry by {created_by.username}",
            auto_created=True
        )
        
        current_time = session['end']
    
    # Final work session (after last break/lunch until check_out)
    if current_time < check_out:
        ActivitySession.objects.create(
            work_day=work_day,
            session_type='work',
            start_time=current_time,
            end_time=check_out,
            notes=f"Auto-created by {created_by.username}",
            auto_created=True
        )
    
    # Calculate totals
    calculate_daily_totals_manual(work_day)
    
    # Add adjustment record
    work_day.add_adjustment_record(
        adjusted_by=created_by,
        reason=reason,
        sessions_affected=[s.id for s in work_day.sessions.all()]
    )
    
    return work_day


def parse_datetime(date_obj, time_str):
    """
    Parse a time string and combine with date
    """
    time_obj = datetime.strptime(time_str, '%H:%M').time()
    return datetime.combine(date_obj, time_obj)
